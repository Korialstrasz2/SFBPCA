"""Archivio delle allerte per la nuova applicazione."""

from __future__ import annotations

import csv
import io
from collections import defaultdict
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


FIELDNAMES = [
    "alert_type",
    "account_id",
    "account_name",
    "contact_id",
    "contact_name",
    "contact_roles",
    "issue_category",
    "data_focus",
    "details",
    "message",
]

EXCEL_COLUMNS: List[Tuple[str, str]] = [
    ("Tipo di allerta", "alert_type"),
    ("Account ID", "account_id"),
    ("Account", "account_name"),
    ("Contatto ID", "contact_id"),
    ("Contatto", "contact_name"),
    ("Ruoli contatto", "contact_roles"),
    ("Focus dati", "data_focus"),
    ("Dettagli", "details"),
    ("Messaggio", "message"),
]


class AlertSummaryStore:
    """Memorizza le allerte prodotte dal ciclo di controllo."""

    def __init__(self) -> None:
        self.reset()

    def reset(self) -> None:
        self._alerts: List[Dict[str, str]] = []
        self._total_accounts: int = 0

    def record(self, alert: Dict[str, str]) -> None:
        if not alert:
            return
        normalised = {field: alert.get(field, "") for field in FIELDNAMES}
        self._alerts.append(normalised)

    def extend(self, alerts: Iterable[Dict[str, str]]) -> None:
        for alert in alerts:
            self.record(alert)

    def all_alerts(self) -> List[Dict[str, str]]:
        return list(self._alerts)

    def summary_rows(self) -> List[Dict[str, str]]:
        return self.all_alerts()

    def statistics(self, total_accounts: Optional[int] = None) -> Dict[str, object]:
        alerts = list(self._alerts)
        if total_accounts is not None:
            self._total_accounts = max(int(total_accounts), 0)
        account_counts: Dict[Tuple[str, str], int] = defaultdict(int)
        account_keys: set[Tuple[str, str]] = set()
        contact_keys: set[Tuple[str, str]] = set()
        per_type: Dict[str, Dict[str, object]] = {}
        alerts_without_contact = 0

        for alert in alerts:
            account_key = (
                (alert.get("account_id") or "").strip(),
                (alert.get("account_name") or "").strip(),
            )
            contact_key = (
                (alert.get("contact_id") or "").strip(),
                (alert.get("contact_name") or "").strip(),
            )
            alert_type = (alert.get("alert_type") or "Sconosciuto").strip() or "Sconosciuto"

            if any(account_key):
                account_counts[account_key] += 1
                account_keys.add(account_key)

            if any(contact_key):
                contact_keys.add(contact_key)
            else:
                alerts_without_contact += 1

            bucket = per_type.setdefault(
                alert_type,
                {
                    "alerts": 0,
                    "accounts": set(),
                    "contacts": set(),
                    "contactless": 0,
                },
            )
            bucket["alerts"] = int(bucket["alerts"]) + 1
            if any(account_key):
                bucket["accounts"].add(account_key)
            if any(contact_key):
                bucket["contacts"].add(contact_key)
            else:
                bucket["contactless"] = int(bucket["contactless"]) + 1

        total_alerts = len(alerts)
        accounts_with_alerts = len(account_keys)
        total_accounts_value = self._total_accounts or accounts_with_alerts
        average_alerts = total_alerts / accounts_with_alerts if accounts_with_alerts else 0.0

        per_type_rows = [
            {
                "alert_type": alert_type,
                "alert_count": int(bucket["alerts"]),
                "unique_accounts": len(bucket["accounts"]),
                "unique_contacts": len(bucket["contacts"]),
                "alerts_without_contact": int(bucket["contactless"]),
            }
            for alert_type, bucket in sorted(per_type.items(), key=lambda item: item[0])
        ]

        top_accounts = [
            {
                "account_id": key[0],
                "account_name": key[1],
                "alert_count": count,
            }
            for key, count in sorted(
                account_counts.items(),
                key=lambda item: item[1],
                reverse=True,
            )
        ]

        top_accounts = top_accounts[:5]

        return {
            "totals": {
                "total_alerts": total_alerts,
                "total_accounts": total_accounts_value,
                "accounts_with_alerts": accounts_with_alerts,
                "unique_contacts": len(contact_keys),
                "unique_alert_types": len(per_type_rows),
                "alerts_without_contact": alerts_without_contact,
                "average_alerts_per_account": round(average_alerts, 2),
            },
            "per_type": per_type_rows,
            "top_accounts": top_accounts,
        }

    def to_csv(self) -> bytes:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=FIELDNAMES)
        writer.writeheader()
        for alert in self._alerts:
            writer.writerow({field: alert.get(field, "") for field in FIELDNAMES})
        return output.getvalue().encode("utf-8")

    def to_excel(self, statistics: Optional[Dict[str, object]] = None) -> bytes:
        workbook = Workbook()
        details_sheet = workbook.active
        details_sheet.title = "Dettagli allerte"

        header_font = Font(bold=True)
        details_sheet.append([header for header, _ in EXCEL_COLUMNS])
        for cell in details_sheet[1]:
            cell.font = header_font

        for alert in self._alerts:
            details_sheet.append([alert.get(field, "") for _, field in EXCEL_COLUMNS])

        detail_widths = [24, 18, 28, 18, 26, 26, 18, 40, 40]
        for index, width in enumerate(detail_widths, start=1):
            details_sheet.column_dimensions[get_column_letter(index)].width = width

        stats_data = statistics or self.statistics()
        stats_sheet = workbook.create_sheet("Statistiche riepilogo")

        stats_sheet.append(["Metrica", "Valore"])
        for cell in stats_sheet[1]:
            cell.font = header_font

        totals = stats_data.get("totals", {}) if isinstance(stats_data, dict) else {}
        metrics = [
            ("Allerte totali", totals.get("total_alerts", 0)),
            ("Account totali analizzati", totals.get("total_accounts", 0)),
            ("Account con allerte", totals.get("accounts_with_alerts", 0)),
            ("Contatti con allerte", totals.get("unique_contacts", 0)),
            ("Tipi di allerta", totals.get("unique_alert_types", 0)),
            ("Allerte senza contatto", totals.get("alerts_without_contact", 0)),
            ("Media allerte per account", totals.get("average_alerts_per_account", 0)),
        ]
        for label, value in metrics:
            stats_sheet.append([label, value])

        stats_sheet.append([])
        stats_sheet.append(["Statistiche per tipo di allerta"])
        stats_sheet.append(
            [
                "Tipo di allerta",
                "Allerte",
                "Account coinvolti",
                "Contatti coinvolti",
                "Allerte senza contatto",
            ]
        )
        for cell in stats_sheet[stats_sheet.max_row]:
            cell.font = header_font

        per_type_rows = stats_data.get("per_type", []) if isinstance(stats_data, dict) else []
        for row in per_type_rows:
            stats_sheet.append(
                [
                    row.get("alert_type", "N/D"),
                    row.get("alert_count", 0),
                    row.get("unique_accounts", 0),
                    row.get("unique_contacts", 0),
                    row.get("alerts_without_contact", 0),
                ]
            )

        stats_sheet.append([])
        stats_sheet.append(["Account con più allerte"])
        stats_sheet.append(["Account", "ID", "Allerte"])
        for cell in stats_sheet[stats_sheet.max_row]:
            cell.font = header_font

        top_accounts = stats_data.get("top_accounts", []) if isinstance(stats_data, dict) else []
        for account in top_accounts:
            label = account.get("account_name") or account.get("account_id") or "Sconosciuto"
            stats_sheet.append([label, account.get("account_id", ""), account.get("alert_count", 0)])

        stats_widths = [36, 24, 24, 24, 24]
        for index, width in enumerate(stats_widths, start=1):
            stats_sheet.column_dimensions[get_column_letter(index)].width = width

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()


ALERT_SUMMARY = AlertSummaryStore()
